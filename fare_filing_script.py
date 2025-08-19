import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font
from openpyxl import load_workbook
from datetime import datetime
import math

#Get the directory in which the file is located
def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return(os.path.dirname(os.path.abspath(__file__)))
    
#Output directory
def resolve_path_output(filename):
    return os.path.join(get_base_dir(),'output',filename)

#Input and data directory
def resolve_path_input(filename):
    return os.path.join(get_base_dir(), 'source', filename)

def is_file_open(file_location):
    try:
        with open(file_location, 'a'):
            return False
    except PermissionError:
        return True

# Mapping for special destination codes
CODE_MAP = {
    'SAW': 'IST', 'OTP': 'BUH', 'BGY': 'MIL',
    'IKA': 'THR', 'GYD': 'BAK', 'ESB': 'ANK', 'VKO' : 'MOW'
}

#Round to the nearest integer
def round_nearest(num):
        return math.floor(num+1)

class FareFilingProcessor:
    def __init__(self, input_path, data_path):
        input_path = resolve_path_input(input_path)
        data_path = resolve_path_input(data_path)

        #Check if input file is open
        if(is_file_open(input_path)):
            input(f"Close the input file.")
            return
        # Read input data
        self.df_table, self.sales, self.travel, self.fn = self.read_input(input_path)
        # Read  data sheet
        (self.fare_class_map,
         self.df_tax,
         self.df_exch,
         self.df_atpco,
         self.df_fod,
         self.df_tfee_discount,
         self.df_restricted_od) = self.read_data(data_path)
        # Invert fare_class_map for RBD->level
        self.inv_fare_map = {v: k for k, v in self.fare_class_map.items()}
        # Prepare output workbook
        self.out_wb = Workbook()
        self.del_ws = self.out_wb.active

        self.del_ws.title = 'DELETE'
        self.del_ws.append(["Tariff","CXR","NAT1","NAT2","LOC1","LOC2","Rule",
                             "FareClass","OW/RT","RTG","FN","CUR","Amount",
                             "Eff.Date","Disc.Date","GFSFAN"])
        
        self.file_ws = self.out_wb.create_sheet('FILE')
        self.file_ws.append(['ACTION','Origin','Dest.','RBD','Channel','OW/RT',
                              'Baggage','Product type','Base Fare','Currency',
                              'SALES','TRAVEL','NOTES','FN','Filing Date',
                              'Total Fare','FBC','DUPE CHECK'])
        
        self.gh_ws = self.out_wb.create_sheet('GH FARE AMENDMENT')
        self.gh_ws.append(["ACTION","Tariff","CXR","NAT1","NAT2","LOC1","LOC2","Rule",
                             "FareClass","OW/RT","RTG","FN","CUR","New Amount",
                             "Eff.Date","Disc.Date","GFSFAN"])

    def read_input(self, input_path):
        df_raw = pd.read_excel(input_path, header=None)
        # SALES row
        sales_row = df_raw[df_raw[0].astype(str).str.strip().str.upper() == 'SALES'].index[0]
        sales = df_raw.iloc[sales_row + 1, 0]
        travel = df_raw.iloc[sales_row + 1, 1]
        fn = df_raw.iloc[sales_row + 1, 2]
        # Header row
        header_row = df_raw[(df_raw[0] == 'O') & (df_raw[1] == 'D')].index[0]
        df_table = df_raw.iloc[header_row + 1:].copy()
        df_table.columns = df_raw.iloc[header_row]
        df_table = df_table.dropna(subset=['O']).reset_index(drop=True)
        # Add COMPLETED column
        cols = list(df_table.columns)
        idx = cols.index('B1') + 1
        df_table.insert(idx, 'COMPLETED', '')
        return df_table, sales, travel, fn
 
    def read_data(self, data_path):
        xls = pd.ExcelFile(data_path)
        df_fcr = pd.read_excel(xls, 'FCR')
        df_tax = pd.read_excel(xls, 'Tax')
        df_exch = pd.read_excel(xls, 'Exchange Rates')
        df_atpco = pd.read_excel(xls, 'ATPCO Data')
        df_fod = pd.read_excel(xls, 'Fare Calc OD')
        fare_class_map = dict(zip(df_fcr['Fare Level'], df_fcr['Fare Class']))
        df_tfee_discount = pd.read_excel(xls, 'Tfee discount')
        df_restricted_od = pd.read_excel(xls, 'Restricted OD')
        return fare_class_map, df_tax, df_exch, df_atpco, df_fod, df_tfee_discount, df_restricted_od
    
    def get_exchange_rate(self, currency):
        # currency is e.g. 'QAR'; lookup 'QAR/AED' in the sheet
        curr = str(currency).strip()
        pair = f"{curr}/AED"
        rates = self.df_exch[self.df_exch['Currency'] == pair]
        if rates.empty:
            raise ValueError(f"No exchange rate found for pair {pair}")
        return rates.iloc[0]['Price']

    def translate_loc(self, code):
        return CODE_MAP.get(code, code)

    def baggage_structure(self, origin, dest, rbd, brand, trip):
        df = self.df_atpco[
            (self.df_atpco['LOC1'].str.strip() == self.translate_loc(origin)) &
            (self.df_atpco['LOC2'].str.strip() == self.translate_loc(dest)) &
            (self.df_atpco['RBD'].str.strip() == rbd) &
            (self.df_atpco['BRAND'].str.strip() == brand) &
            (self.df_atpco['OW/RT'] == trip)
        ]
        return (df['BAG'].values[0]) if not df.empty else 0
    
    def baggage_non_structure(self, origin, dest, rbd, brand, fn, trip):
        df = self.df_atpco[
            (self.df_atpco['LOC1'].str.strip() == self.translate_loc(origin)) &
            (self.df_atpco['LOC2'].str.strip() == self.translate_loc(dest)) &
            (self.df_atpco['RBD'].str.strip() == rbd) &
            (self.df_atpco['BRAND'].str.strip() == brand) &
            (self.df_atpco['FN'].str.strip() == fn) &
            (self.df_atpco['OW/RT'] == trip)
        ]
        return (df['BAG'].values[0]) if not df.empty else 0

    def get_new_rbd(self, amount, trip):
        amount = int(amount)
        if trip == 1:
            thresholds = [
                (50,389,'L'),(390,454,'Q'),(455,532,'H'),(533,623,'K'),
                (624,714,'U'),(715,831,'B'),(832,987,'R'),(988,1182,'N'),
                (1183,1377,'M'),(1378,1637,'T'),(1638,1962,'W'),(1963,2352,'O'),
                (2353,2807,'E'),(2808,3262,'I'),(3263,3782,'A'),(3783,999999999999999,'Y') 
            ]
        
        elif trip == 2:
            thresholds = [
                (100,599,'L'),(600,699,'Q'),(700,819,'H'),(820,959,'K'),
                (960,1099,'U'),(1100,1279,'B'),(1280,1519,'R'),(1520,1819,'N'),
                (1820,2119,'M'),(2120,2519,'T'),(2520,3019,'W'),(3020,3619,'O'),
                (3620,4319,'E'),(4320,5019,'I'),(5020,5819,'A'),(5820,999999999999999,'Y')
            ]
    
        for low, high, rbd in thresholds:
            if low <= amount <= high:
                return rbd

    def get_baggage_code(self,bag):
        if(bag==20):
            return "B"
        elif(bag==30):
            return "L"
        elif(bag==40):
            return "X"
        elif(bag>40):
            return "NF"

    def fbc_calc(self, origin, destination, trip, brand, channel, sales, fn, rbd, bag_code):
        trip_code = 'O' if trip == 1 else 'R'
        sales_u = str(sales).strip().upper()
        # Brand code
        if sales_u == 'STRUCTURE':
            if brand == 'Brand 1': code = '6'
            elif brand in ('Brand 2', 'GDS 1'): code = '7'
            elif brand == 'Brand 3': code = '7'
            elif brand == 'GDS 2': code = '3'
        else:
            if brand == 'Brand 1': code = '6'
            elif brand == 'Brand 2': code = '7'
            elif brand == 'Brand 3': code = '8'
            elif brand == 'GDS 1': code = 'P7'
            elif brand == 'GDS 2': code = 'P3'
        # Origin country
        country_map = {'BAH':'BH','KWI':'KW','DOH':'QA','MCT':'OM','SLL':'OM'}
        o_country = country_map.get(origin, 'SA')
        # Type
        type = '2' if channel == 'WEB' and sales_u == 'STRUCTURE' else ('5' if channel == 'WEB' else '1')
        return f"{origin}{destination}{rbd}{trip_code}{bag_code}{code}{o_country}{type}-{fn}"

    def write_file(self, action, origin, dest, rbd, channel, trip,
                   baggage, brand, base_fare, currency,
                   total_fare, fbc, notes=''):

        self.base_fare = base_fare
        self.total_fare = total_fare

        if(base_fare != int(base_fare)):
            self.base_fare = round_nearest(base_fare)
        
        if(total_fare != int(total_fare)):
            self.total_fare = round_nearest(total_fare)
        
        if(self.action == "AMEND"):
            self.base_fare,self.total_fare = self.amend_same_fare(brand, self.base_fare, self.total_fare)

        row = [action, origin, dest, rbd, channel, trip,
               baggage, brand, self.base_fare, currency,
               self.sales, self.travel, notes, self.fn,
               datetime.now().strftime('%d-%m-%y'),
               self.total_fare, fbc, '']
        self.file_ws.append(row)

    def write_del(self, origin, dest, rbd, brand, trip):
        df = self.df_atpco[
            (self.df_atpco['LOC1'].str.strip() == self.translate_loc(origin)) &
            (self.df_atpco['LOC2'].str.strip() == self.translate_loc(dest)) &
            (self.df_atpco['RBD'].str.strip() == rbd) &
            (self.df_atpco['BRAND'].str.strip() == brand) &
            (self.df_atpco['FN'].str.strip() == self.fn) &
            (self.df_atpco['OW/RT'] == trip)
        ]
        if not df.empty:
            row = df.iloc[0]
            out = [row['Tariff'], row['CXR'], row['NAT1'], row['NAT2'],
                   row['LOC1'], row['LOC2'], row['Rule'], row['FareClass'],
                   row['OW/RT'], row['RTG'], row['FN'], row['CUR'], row['Amount'],
                   row['Eff.Date'].strftime("%d/%m/%y"), row['Disc.Date'], row['GFSFAN']]
            self.del_ws.append(out)

    def amend_same_fare(self, brand, base_fare, total_fare):
        df = self.df_atpco[
            (self.df_atpco['LOC1'] == self.translate_loc(self.origin)) &
            (self.df_atpco['LOC2'] == self.translate_loc(self.dest)) &
            (self.df_atpco['RBD'] == self.fare_class_map[self.filed_level]) &
            (self.df_atpco['BRAND'] == brand) &
            (self.df_atpco['FN'] == self.fn) &
            (self.df_atpco['OW/RT'] == self.trip)
        ]
        
        # print(base_fare, total_fare, "Before the condition") 
        if(not df['BASE FARE'].empty and (df['BASE FARE'].item() - base_fare)==0):
            
            base_fare+=1; total_fare+=1
            
            # print(base_fare, total_fare, "After the condition in BF", df['BASE FARE'].item())

        elif(not df['TOTAL FARE'].empty and (df['BASE FARE'].item() - base_fare)==0):
            
            base_fare+=1
            total_fare+=1
            # print(base_fare, total_fare,"After the condition in TF", df['BASE FARE'].item())

        return base_fare, total_fare

    def brand1_calc(self):
        brand = 'Brand 1'; channel = 'WEB'
        self.filed_rbd = self.fare_class_map[self.filed_level]
        if(self.trip == 1):
            if(self.filed_rbd == "L"):
                self.b1_base_fare_with_yq_aed = 325
            elif(self.filed_rbd == "Q"):
                self.b1_base_fare_with_yq_aed = 390
            elif(self.filed_rbd == "H"):
                self.b1_base_fare_with_yq_aed = 455
            elif(self.filed_rbd == "K"):
                self.b1_base_fare_with_yq_aed = 533
            elif(self.filed_rbd == "U"):
                self.b1_base_fare_with_yq_aed = 624
            elif(self.filed_rbd == "B"):
                self.b1_base_fare_with_yq_aed = 715
            elif(self.filed_rbd == "R"):
                self.b1_base_fare_with_yq_aed = 832
            elif(self.filed_rbd == "N"):
                self.b1_base_fare_with_yq_aed = 988
            elif(self.filed_rbd == "M"):
                self.b1_base_fare_with_yq_aed = 1183
            elif(self.filed_rbd == "T"):
                self.b1_base_fare_with_yq_aed = 1378 
            elif(self.filed_rbd == "W"):
                self.b1_base_fare_with_yq_aed = 1638
            elif(self.filed_rbd == "Y"):
                self.b1_base_fare_with_yq_aed = 1963
            elif(self.filed_rbd == "E"):
                self.b1_base_fare_with_yq_aed = 2353
            elif(self.filed_rbd == "I"):
                self.b1_base_fare_with_yq_aed = 2808
            elif(self.filed_rbd == "A"):
                self.b1_base_fare_with_yq_aed = 3263  
            elif(self.filed_rbd == "Y"):
                self.b1_base_fare_with_yq_aed = 3783 
        if(self.trip == 2):
            if(self.filed_rbd == "L"):
                self.b1_base_fare_with_yq_aed = 500
            elif(self.filed_rbd == "Q"):
                self.b1_base_fare_with_yq_aed = 600
            elif(self.filed_rbd == "H"):
                self.b1_base_fare_with_yq_aed = 700
            elif(self.filed_rbd == "K"):
                self.b1_base_fare_with_yq_aed = 820
            elif(self.filed_rbd == "U"):
                self.b1_base_fare_with_yq_aed = 960
            elif(self.filed_rbd == "B"):
                self.b1_base_fare_with_yq_aed = 1100
            elif(self.filed_rbd == "R"):
                self.b1_base_fare_with_yq_aed = 1280
            elif(self.filed_rbd == "N"):
                self.b1_base_fare_with_yq_aed = 1520
            elif(self.filed_rbd == "M"):
                self.b1_base_fare_with_yq_aed = 1820
            elif(self.filed_rbd == "T"):
                self.b1_base_fare_with_yq_aed = 2120 
            elif(self.filed_rbd == "W"):
                self.b1_base_fare_with_yq_aed = 2520
            elif(self.filed_rbd == "Y"):
                self.b1_base_fare_with_yq_aed = 3020
            elif(self.filed_rbd == "E"):
                self.b1_base_fare_with_yq_aed = 3620
            elif(self.filed_rbd == "I"):
                self.b1_base_fare_with_yq_aed = 4320
            elif(self.filed_rbd == "A"):
                self.b1_base_fare_with_yq_aed = 5020  
            elif(self.filed_rbd == "Y"):
                self.b1_base_fare_with_yq_aed = 5820

        self.b1_base_fare = (self.b1_base_fare_with_yq_aed/self.exch) - self.yq_tax
        
        self.b1_total_fare = self.b1_base_fare + self.tax + self.yq_tax

        
        fbc = self.fbc_calc(self.origin, self.dest, self.trip, brand, channel, self.sales, self.fn, self.filed_rbd, "")
        self.write_file(self.action, self.origin, self.dest, self.fare_class_map[self.filed_level],
                        channel, self.trip, self.b1_baggage,
                        brand, self.b1_base_fare, self.currency,
                        self.b1_total_fare, fbc)

    def brand2_calc(self):
        brand = 'Brand 2'; channel = 'WEB';
        diff = 0
        b1_total_fare_aed = self.b1_total_fare * self.exch
        if b1_total_fare_aed <= 500:
            diff = 20 if self.trip == 1 else 40
        elif b1_total_fare_aed <= 1000:
            diff = 30 if self.trip == 1 else 60
        elif b1_total_fare_aed <= 1500:
            diff = 40 if self.trip == 1 else 80
        elif b1_total_fare_aed <= 2000:
            diff = 50 if self.trip == 1 else 100
        else:
            diff = 80 if self.trip == 1 else 160
        
        diff = diff/self.exch
        self.b2_total_fare = self.b1_total_fare + diff

        self.b2_base_fare = self.b2_total_fare - self.tax - self.yq_tax
       
        bag_code = self.get_baggage_code(self.b2_baggage)
        fbc = self.fbc_calc(self.origin, self.dest, self.trip, brand, channel, self.sales, self.fn, self.filed_rbd,bag_code)
        self.write_file(self.action, self.origin, self.dest, self.fare_class_map[self.filed_level],
                        channel, self.trip, self.b2_baggage,
                        brand, self.b2_base_fare, self.currency,
                        self.b2_total_fare, fbc)

    def gds1_calc(self):
        brand = 'GDS 1'; channel = 'GDS';
        
        # segment fee
        b2_total_fare_aed = self.b2_total_fare * self.exch
        if b2_total_fare_aed <= 500:
            seg_fee = 20 if self.trip == 1 else 40
        elif b2_total_fare_aed <= 1000:
            seg_fee = 30 if self.trip == 1 else 60
        elif b2_total_fare_aed <= 1500:
            seg_fee = 40 if self.trip == 1 else 80
        elif b2_total_fare_aed <= 2000:
            seg_fee = 50 if self.trip == 1 else 100
        else:
            seg_fee = 80 if self.trip == 1 else 160

        
        self.segment_fee = seg_fee/self.exch      
        if(self.filed_level<=5):
            origin_count = 0 
            dest_count = 0
            for origin in self.df_fod['Origin']:
                if(type(origin)==str):
                    if(self.origin == origin.strip()):
                        origin_count+=1

            for dest in self.df_fod['Destination']:
                if(type(dest) == str):
                    if(self.dest == dest.strip()):
                        dest_count+=1

            if(origin_count>0 and dest_count>0):
                self.gds1_base_fare = self.b2_base_fare + self.segment_fee - self.tfee
            else:
                self.gds1_base_fare = self.b2_base_fare + self.segment_fee

        else:
            self.gds1_base_fare = self.b2_base_fare + self.segment_fee
            
        self.gds1_total_fare = self.gds1_base_fare + self.tax + self.yq_tax + self.tfee

        bag_code = self.get_baggage_code(self.gds1_baggage)
        fbc = self.fbc_calc(self.origin, self.dest, self.trip, brand, channel, self.sales, self.fn, self.filed_rbd,bag_code)
        self.write_file(self.action, self.origin, self.dest, self.fare_class_map[self.filed_level],
                        channel, self.trip, self.gds1_baggage,
                        brand, self.gds1_base_fare, self.currency,
                        self.gds1_total_fare, fbc)

    def brand3_calc(self):
        brand = 'Brand 3'; channel = 'WEB';
        if self.trip == 1:
            self.b3_total_fare = self.gds1_total_fare + 100/self.exch
        else:
            self.b3_total_fare = self.gds1_total_fare + 200/self.exch

        self.b3_base_fare = self.b3_total_fare - self.yq_tax - self.tax

        bag_code = self.get_baggage_code(self.b3_baggage)
        fbc = self.fbc_calc(self.origin, self.dest, self.trip, brand, channel, self.sales, self.fn, self.filed_rbd,bag_code)
        self.write_file(self.action, self.origin, self.dest, self.fare_class_map[self.filed_level],
                        channel, self.trip, self.b3_baggage,
                        brand, self.b3_base_fare, self.currency,
                        self.b3_total_fare, fbc)

    def gds2_calc(self):
        brand = 'GDS 2'; channel = 'GDS';
        level = self.inv_fare_map[self.filed_rbd]
        flex = 0.05 if 1 <= level <= 13 else 0.1
        self.gds2_base_fare = self.b3_base_fare + (flex * (self.b3_base_fare + self.yq_tax)) + self.segment_fee

        self.gds2_total_fare = self.gds2_base_fare + self.yq_tax + self.tax + self.tfee

        bag_code = self.get_baggage_code(self.gds2_baggage)
        fbc = self.fbc_calc(self.origin, self.dest, self.trip, brand, channel, self.sales, self.fn, self.filed_rbd,bag_code)
        self.write_file(self.action, self.origin, self.dest, self.fare_class_map[self.filed_level],
                        channel, self.trip, self.gds2_baggage,
                        brand, self.gds2_base_fare, self.currency,
                        self.gds2_total_fare, fbc)

    def write_gh(self, df):
        row = df.iloc[0]
        # self.gh_ws.cell(row, 14).font = Font(color = "FF0000")
        if(self.new_base_fare != int(self.new_base_fare)):
            self.new_base_fare = round_nearest(self.new_base_fare)
        out = [self.action, row['Tariff'], row['CXR'], row['NAT1'], row['NAT2'],
                row['LOC1'], row['LOC2'], row['Rule'], row['FareClass'],
                row['OW/RT'], row['RTG'], row['FN'], row['CUR'], self.new_base_fare,
                row['Eff.Date'].strftime("%d/%m/%y"), row['Disc.Date'], row['GFSFAN']]
        self.gh_ws.append(out)

    def gh_lookup(self,brand):    
        df = self.df_atpco[
            (self.df_atpco['LOC1'] == self.translate_loc(self.origin)) &
            (self.df_atpco['LOC2'] == self.translate_loc(self.dest)) &
            (self.df_atpco['RBD'] == "GH") &
            (self.df_atpco['BRAND'] == brand) &
            (self.df_atpco['FN'] == self.fn) &
            (self.df_atpco['OW/RT'] == self.trip)
        ]
       
        return df

    def gh_calc(self):
        self.action = "Amend Fare"
        self.channel = "TA"
        if(self.trip == 1):
            if(self.currency == "QAR" or self.currency == "SAR"):
                gh_increment = 20
            else:
                gh_increment = 2
        else:
            if(self.currency == "QAR" or self.currency == "SAR"):
                gh_increment = 40
            else:
                gh_increment = 4

        brand = "Brand 1"
        df = self.gh_lookup(brand)  
        if(not df.empty):
            self.new_base_fare = self.b1_base_fare + gh_increment
            self.write_gh(df)

        brand = "Brand 2"
        df = self.gh_lookup(brand)  
        if(not df.empty):
            self.new_base_fare = self.b2_base_fare + gh_increment
            self.write_gh(df)

        brand = "Brand 3"
        df = self.gh_lookup(brand)  
        if(not df.empty):
            self.new_base_fare = self.b3_base_fare + gh_increment
            self.write_gh(df)
            
    def amend(self):
        self.action = 'AMEND'
        self.filed_rbd = self.fare_class_map[self.filed_level]
        if(self.filed_level>8):
            return
        # Check current base fare in ATPCO
        df = self.df_atpco[
            (self.df_atpco['LOC1'] == self.translate_loc(self.origin)) &
            (self.df_atpco['LOC2'] == self.translate_loc(self.dest)) &
            (self.df_atpco['RBD'] == self.fare_class_map[self.filed_level]) &
            (self.df_atpco['BRAND'] == 'Brand 1') &
            (self.df_atpco['FN'] == self.fn) &
            (self.df_atpco['OW/RT'] == self.trip)
        ]
        if not df.empty and ((df.iloc[0]['BASE FARE']) - self.b1_base_fare) == 0:
               
                existing_content = self.df_table.at[self.idx, 'COMPLETED']
                new_content = str(existing_content) + '//Not amended as same fare'
                self.df_table.at[self.idx, 'COMPLETED'] = new_content.strip()
                return
        if df.empty:
            return
        
        self.channel = "WEB"
        
        fbc = self.fbc_calc(self.origin, self.dest, self.trip, "Brand 1", self.channel, self.sales, self.fn, self.filed_rbd,"")
        self.write_file(self.action, self.origin, self.dest, self.fare_class_map[self.filed_level],
                        self.channel, self.trip, self.b1_baggage,
                        "Brand 1", self.b1_base_fare, self.currency,
                        self.b1_total_fare, fbc)
        self.b1_base_fare = self.base_fare
        self.b1_total_fare = self.total_fare
        self.brand2_calc()
        self.b2_base_fare = self.base_fare
        self.b2_total_fare = self.total_fare
        self.gds1_calc()
        self.gds1_base_fare = self.base_fare
        self.gds1_total_fare = self.total_fare
        self.brand3_calc()
        self.b3_base_fare = self.base_fare
        self.b3_total_fare = self.total_fare
        self.gds2_calc()
        self.gds2_base_fare = self.base_fare
        self.gds2_total_fare = self.total_fare
        self.df_table.at[self.idx, 'COMPLETED'] = 'YES'
        self.gh_calc()              

    def build(self):
        self.action = 'NEW'
        # use current filed_level to set RBD
        self.filed_rbd = self.fare_class_map[self.filed_level]
        self.brand1_calc()
        self.brand2_calc()
        self.gds1_calc()
        self.brand3_calc()
        self.gds2_calc()

    def delete(self):
        self.action = 'DEL'
        if(self.filed_level>8):
            return
        for brand in ['Brand 1','Brand 2','Brand 3','GDS 1','GDS 2']:
            self.write_del(self.origin, self.dest,
                           self.fare_class_map[self.filed_level], brand, self.trip)

    def error_check(self):
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        seen = set()
        for row in range(2, self.file_ws.max_row + 1):
            fbc = self.file_ws.cell(row, 17).value
            if fbc in seen:
                self.file_ws.cell(row, 17).fill = red_fill
                self.file_ws.cell(row, 18, value='Not OK')
            else:
                seen.add(fbc)
                self.file_ws.cell(row, 18, value='OK')

    def process(self):
        for self.idx, row in self.df_table.iterrows():
            self.origin = row['O']
            self.dest = row['D']
            self.trip = row['O/R']
            self.filed_rbd = row['RBD']
            self.currency = row['CURRENCY']
            self.b1_total_fare = row['B1']
            
            
            #Check if input data is blank
            if pd.isna(self.dest) or pd.isna(self.trip) or pd.isna(self.filed_rbd) or pd.isna(self.currency) or pd.isna(self.b1_total_fare):
                self.df_table.at[self.idx, 'COMPLETED'] = 'Missing input data'
                continue
        
            if(type(self.b1_total_fare)!=int and type(self.b1_total_fare)!= float):
                self.df_table.at[self.idx, 'COMPLETED'] = 'Incorrect B1 fare in input sheet'
                continue

            final_total_fare = self.b1_total_fare

            #Check if origin and destination is in the restricted list

            res_origin_count = 0 
            res_dest_count = 0

            for res_origin in self.df_restricted_od['Origin']:
                if(self.origin == res_origin.strip()):
                    res_origin_count+=1
            
            for res_dest in self.df_restricted_od['Destination']:
                if(self.dest == res_dest.strip()):
                    res_dest_count+=1

            if(res_origin_count>0 and res_dest_count>0):
                self.df_table.at[self.idx, 'COMPLETED'] = 'Restricted OD'
                continue            
            
            #Check if correct origin is in the input sheet
            origin_count = 0
            for origin_cell_value in self.df_fod["Origin"]:
                if(type(origin_cell_value) == str):
                    if(self.origin == origin_cell_value.strip() or self.origin == "SLL"):
                        origin_count+=1
            
            if(origin_count<=0):
                self.df_table.at[self.idx, 'COMPLETED'] = 'Incorrect origin in input sheet'
                continue
            
            #Check if correct destination is in the input sheet
            dest_count = 0
            for dest_cell_value in self.df_fod["All Destination"]:
                if(type(dest_cell_value) == str):
                    if(self.dest == dest_cell_value.strip()):
                        dest_count+=1
            
            if(dest_count<=0):
                self.df_table.at[self.idx, 'COMPLETED'] = 'Incorrect destination in input sheet'
                continue

            #Check if incorrect trip type is in the input sheet
            if(self.trip!=1 and self.trip!=2):
                self.df_table.at[self.idx, 'COMPLETED'] = 'Incorrect trip type in input sheet'
                continue
            
            if(type(self.filed_rbd) == str):
                if(len(self.filed_rbd)!=1):
                    self.df_table.at[self.idx, 'COMPLETED'] = 'Incorrect RBD in input sheet'
                    continue
            else:
                self.df_table.at[self.idx, 'COMPLETED'] = 'Incorrect RBD in input sheet'  
                continue

            #Check if incorrect currency is in the input sheet
            if(self.currency!="BHD" and self.currency!="KWD" and self.currency!="QAR" and self.currency!="SAR" and self.currency!="OMR"):
                self.df_table.at[self.idx, 'COMPLETED'] = 'Incorrect currency in input sheet'
                continue

            self.filed_level = self.inv_fare_map[self.filed_rbd]
            # Calculate Baggage
            if(self.filed_level>8):
                self.b2_baggage = self.baggage_structure(self.origin, self.dest, self.filed_rbd, 'Brand 2',self.trip)
                self.b3_baggage = self.baggage_structure(self.origin, self.dest, self.filed_rbd, 'Brand 3',self.trip)
                self.gds1_baggage = self.baggage_structure(self.origin, self.dest, self.filed_rbd, 'GDS 1',self.trip)
                self.gds2_baggage = self.baggage_structure(self.origin, self.dest, self.filed_rbd, 'GDS 2',self.trip)
                self.b1_baggage = self.baggage_structure(self.origin, self.dest, self.filed_rbd, 'Brand 1',self.trip)
            else:
                self.b2_baggage = self.baggage_non_structure(self.origin, self.dest, self.filed_rbd, 'Brand 2', self.fn, self.trip)
                self.b3_baggage = self.baggage_non_structure(self.origin, self.dest, self.filed_rbd, 'Brand 3', self.fn, self.trip)
                self.gds1_baggage = self.baggage_non_structure(self.origin, self.dest, self.filed_rbd, 'GDS 1', self.fn, self.trip)
                self.gds2_baggage = self.baggage_non_structure(self.origin, self.dest, self.filed_rbd, 'GDS 2', self.fn, self.trip)
                self.b1_baggage = self.baggage_non_structure(self.origin, self.dest, self.filed_rbd, 'Brand 1', self.fn, self.trip)

            # Check if baggage data is available in the ATPCO
            if(self.b1_baggage==0 or self.b2_baggage==0 or self.b3_baggage==0 or self.gds1_baggage==0 or self.gds2_baggage==0):
                self.df_table.at[self.idx, 'COMPLETED'] = 'Missing ATPCO data'
                continue
                            
            # Calcualte taxes & fees
            if(self.trip == 1):
                self.trip_tax = "OW"
            elif (self.trip == 2):
                self.trip_tax = "RT"

            tax_data = self.df_tax[(self.df_tax['Origin'].str.strip()==self.origin) &
                               (self.df_tax['Destination'].str.strip()==self.dest) &
                               (self.df_tax['JourneyType'].str.strip()==self.trip_tax)].iloc[0]
            self.tax = tax_data['FixedTaxTotal']
            self.yq_tax = tax_data['YQ']
            self.tfee = tax_data['YR']

            #Check if data in tax sheet is blank // if not round to the nearest integer
            if pd.isna(self.tax) or pd.isna(self.yq_tax) or pd.isna(self.tfee):
                self.df_table.at[self.idx, 'COMPLETED'] = 'Missing Tax data'
                continue

            #Calculate base fare with yq in AED to get the fare level
            self.b1_base_fare = self.b1_total_fare - self.tax - self.yq_tax
            final_base_fare = self.b1_base_fare
            self.b1_base_fare_with_yq = self.b1_total_fare - self.tax
            self.exch = self.get_exchange_rate(self.currency)
            self.b1_base_fare_with_yq_aed = self.b1_base_fare_with_yq * self.exch
            self.b1_base_fare_with_yq_aed = self.b1_base_fare_with_yq_aed.astype(int)
            
            if(self.trip == 1):
                if(self.b1_base_fare_with_yq_aed<50):
                    self.df_table.at[self.idx, 'COMPLETED'] = 'Incorrect B1 total fare'
                    continue
            if(self.trip ==2):
                if(self.b1_base_fare_with_yq_aed<50):
                    self.df_table.at[self.idx, 'COMPLETED'] = 'Incorrect B1 total fare'
                    continue

            # Determine new RBD and levels
            self.new_rbd = self.get_new_rbd(self.b1_base_fare_with_yq_aed, self.trip)
            self.new_level = self.inv_fare_map[self.new_rbd]
            #Calculate tfee discount
            od = self.origin + self.dest
            if(self.filed_level <=5):
                filtered_tfee_row = self.df_tfee_discount[(self.df_tfee_discount['Ods'].str.strip()==od)]
                if not filtered_tfee_row.empty:
                    tfee_row = filtered_tfee_row.iloc[0]
                    if(self.trip == 1):
                        self.tfee = tfee_row["OW"]
            
                    if(self.trip == 2):
                        self.tfee = tfee_row["RT"]
                #Check if tfee discount row is empty // else round to the nearest integer
                if pd.isna(self.tfee):
                    self.df_table.at[self.idx, 'COMPLETED'] = 'Missing TFEE data in Fare Calc OD sheet'
                    continue

            if(self.filed_level > 9):
                self.filed_level = 9

            self.brand = "Brand 1"
            
            if (self.filed_level == self.new_level):
                self.amend()
            elif (self.filed_level > self.new_level):
                while (self.filed_level - self.new_level) > 0:
                    self.filed_level -= 1
                    if(self.filed_level - self.new_level == 0):
                        self.channel = "WEB"
                        self.action = "NEW"
                        self.b1_total_fare = final_total_fare
                        self.b1_base_fare = final_base_fare
                        self.filed_rbd = self.fare_class_map[self.filed_level]
                        fbc = self.fbc_calc(self.origin, self.dest, self.trip, "Brand 1", self.channel, self.sales, self.fn, self.filed_rbd,"")
                        self.write_file(self.action, self.origin, self.dest, self.fare_class_map[self.filed_level],
                        self.channel, self.trip, self.b1_baggage,
                        "Brand 1", self.b1_base_fare, self.currency,
                        self.b1_total_fare, fbc)
                        self.brand2_calc()
                        self.gds1_calc()
                        self.brand3_calc()
                        self.gds2_calc()
                    else:
                        self.build()
                
                self.gh_calc()
                existing_content = self.df_table.at[self.idx, 'COMPLETED']
                new_content = str(existing_content) + ' YES'
                self.df_table.at[self.idx, 'COMPLETED'] = new_content.strip()
            else:
                while (self.filed_level <= self.new_level):
                    if (self.filed_level == self.new_level):
                        df = self.df_atpco[
                            (self.df_atpco['LOC1'].str.strip() == self.translate_loc(self.origin)) &
                            (self.df_atpco['LOC2'].str.strip() == self.translate_loc(self.dest)) &
                            (self.df_atpco['RBD'].str.strip() == self.fare_class_map[self.filed_level]) &
                            (self.df_atpco['BRAND'].str.strip() == 'Brand 1') &
                            (self.df_atpco['FN'].str.strip() == self.fn) &
                            (self.df_atpco['OW/RT'] == self.trip)
                        ]              
                        if not df.empty and -1 <= (df.iloc[0]['BASE FARE'] - final_base_fare) <= 1:
                            if(self.currency == "SAR" or self.currency == "QAR"):
                                final_base_fare+=10
                                final_total_fare+=10
                            else:
                                final_base_fare+=1
                                final_total_fare+=1

                            
                            self.b1_total_fare = final_total_fare
                            self.b1_base_fare = final_base_fare
                            existing_content = self.df_table.at[self.idx, 'COMPLETED']
                            new_content = str(existing_content) + ' YES'
                            self.df_table.at[self.idx, 'COMPLETED'] = new_content.strip()
                            self.amend()
                              
                        else:
                            existing_content = self.df_table.at[self.idx, 'COMPLETED']
                            new_content = str(existing_content) + ' YES'
                            self.df_table.at[self.idx, 'COMPLETED'] = new_content.strip()
                            self.amend()                                                       
                    else:
                        self.delete()
                    self.filed_level+=1
            if(self.filed_level>8 or self.new_level>8):
                existing_content = self.df_table.at[self.idx, 'COMPLETED']
                new_content = str(existing_content) + '//Structure RBD'
                self.df_table.at[self.idx, 'COMPLETED'] = new_content.strip()
        path = resolve_path_input('input.xlsx')
        with pd.ExcelWriter(path,
                            engine='openpyxl',
                            mode='a',
                            if_sheet_exists='replace') as writer:
                                # overwrite the “Processed” sheet if it already exists
                                self.df_table.to_excel(writer, sheet_name='Processed', index=False)

        #Final duplicate check and save
        self.error_check()
        output_path = resolve_path_output('output.xlsx')
        if(is_file_open(output_path)):
            input("Close the output file")
            return
        self.out_wb.save(output_path)
        print(f"Output written to {output_path}")
        input("\nPress Enter to exit...")


if __name__ == '__main__':
    print("Filing script is running...")
    processor = FareFilingProcessor('input.xlsx', 'data.xlsx')
    processor.process()
